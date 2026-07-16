from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django.db.models import Q, Prefetch
from apps.core.models import Roles, UserMaster, UserRoles
from apps.fee.models import FeeTemplate, StudentFeeAssignment, FeeConcession
from apps.school.models import School
from apps.school.models.school import AcademicYear, Grade, Section, Student, StudentDocument, Staff, StaffDocument, \
    Branch, Subject
from shared.enums.roles import RolesEnum
from shared.helpers.rbac import check_permission
from shared.helpers.student import get_or_create_parent
from shared.mixins import CustomResponse, CustomPageNumberPagination
from shared.permissions import HasPermission
from openpyxl import load_workbook
from django.conf import settings

from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django.db import IntegrityError
from apps.core.models import (
    Roles,
    UserRoles,
)
from shared.permissions.rbac import HasPermission
from shared.enums.roles import RolesEnum
from django.db import transaction

from shared.utils.fee import generate_student_fees
from io import BytesIO

from openpyxl import Workbook

from django.http import HttpResponse, FileResponse

from rest_framework.views import APIView

from rest_framework.permissions import IsAuthenticated

from shared.utils.logger import audit_logger, application_logger


class CreateAcademicYearAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "academic_year.create"

    def post(self, request):

        school = request.school

        audit_logger.info(
            "academic_year_create_started",
            performed_by=str(request.user.id),
            school_id=str(school.id) if school else None,
        )

        if school is None:
            audit_logger.warning(
                "academic_year_create_failed",
                performed_by=str(request.user.id),
                reason="school_not_found",
            )
            return CustomResponse.errorResponse(description="School not found.")

        academic_year = AcademicYear.objects.create(
            school=school,
            name=request.data.get("name"),
            start_date=request.data.get("start_date"),
            end_date=request.data.get("end_date"),
            status=request.data.get("status", "ACTIVE"),
        )

        audit_logger.info(
            "academic_year_created",
            performed_by=str(request.user.id),
            academic_year_id=str(academic_year.id),
            school_id=str(school.id),
            name=academic_year.name,
            status=academic_year.status,
        )

        return CustomResponse.successResponse(
            description="Academic year created successfully",
            data={"id": academic_year.id},
        )


class AcademicYearListAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "academic_year.view"

    def get(self, request):

        school = request.school

        application_logger.info(
            "academic_years_fetch_started",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
        )

        queryset = AcademicYear.objects.filter(school=school).order_by("-created_at")

        data = [
            {
                "id": obj.id,
                "name": obj.name,
                "start_date": obj.start_date,
                "end_date": obj.end_date,
                "status": obj.status,
            }
            for obj in queryset
        ]

        application_logger.info(
            "academic_years_fetched",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            total_count=len(data),
        )

        return CustomResponse.successResponse(data=data)



class UpdateAcademicYearAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "academic_year.update"

    def put(self, request, academic_year_id):

        school = request.school

        audit_logger.info(
            "academic_year_update_started",
            performed_by=str(request.user.id),
            academic_year_id=str(academic_year_id),
            school_id=str(school.id) if school else None,
        )

        academic_year = AcademicYear.objects.filter(id=academic_year_id, school=school).first()

        if academic_year is None:
            audit_logger.warning(
                "academic_year_update_failed",
                performed_by=str(request.user.id),
                academic_year_id=str(academic_year_id),
                school_id=str(school.id) if school else None,
                reason="academic_year_not_found",
            )
            return CustomResponse.errorResponse(description="Academic year not found.")

        old_status = academic_year.status

        academic_year.name = request.data.get("name", academic_year.name)
        academic_year.start_date = request.data.get("start_date", academic_year.start_date)
        academic_year.end_date = request.data.get("end_date", academic_year.end_date)
        academic_year.status = request.data.get("status", academic_year.status)
        academic_year.save()

        audit_logger.info(
            "academic_year_updated",
            performed_by=str(request.user.id),
            academic_year_id=str(academic_year.id),
            school_id=str(school.id),
            name=academic_year.name,
            old_status=old_status,
            new_status=academic_year.status,
        )

        return CustomResponse.successResponse(description="Academic year updated successfully")

class CreateGradeAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "grade.create"

    def post(self, request):

        school = request.school

        audit_logger.info(
            "grade_create_started",
            performed_by=str(request.user.id),
            school_id=str(school.id) if school else None,
        )

        academic_year = AcademicYear.objects.filter(id=request.data.get("academic_year_id"), school=school).first()

        if academic_year is None:
            audit_logger.warning(
                "grade_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id) if school else None,
                academic_year_id=request.data.get("academic_year_id"),
                reason="academic_year_not_found",
            )
            return CustomResponse.errorResponse(description="Academic Year not found.")

        grade_name = request.data.get("name")

        if Grade.objects.filter(school=school, academic_year=academic_year, name=grade_name).exists():
            audit_logger.warning(
                "grade_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                academic_year_id=str(academic_year.id),
                grade_name=grade_name,
                reason="grade_already_exists",
            )
            return CustomResponse.errorResponse(description="Grade already exists.")

        grade = Grade.objects.create(
            school=school,
            academic_year=academic_year,
            name=grade_name,
            display_order=request.data.get("display_order"),
            status=request.data.get("status", Grade.Status.ACTIVE),
        )

        audit_logger.info(
            "grade_created",
            performed_by=str(request.user.id),
            grade_id=str(grade.id),
            school_id=str(school.id),
            academic_year_id=str(academic_year.id),
            grade_name=grade.name,
            status=grade.status,
        )

        return CustomResponse.successResponse(
            description="Grade created successfully.",
            data={"id": str(grade.id)},
        )

class GradeListAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "grade.view"

    def get(self, request):

        school = request.school

        application_logger.info(
            "grades_fetch_started",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
        )

        grades = Grade.objects.select_related("school", "academic_year").filter(school=school).order_by("display_order")

        data = [
            {
                "id": str(grade.id),
                "school": grade.school.name,
                "academic_year": grade.academic_year.name,
                "name": grade.name,
                "display_order": grade.display_order,
                "status": grade.status,
            }
            for grade in grades
        ]

        application_logger.info(
            "grades_fetched",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            total_count=len(data),
        )

        return CustomResponse.successResponse(data=data)


class UpdateGradeAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "grade.update"

    def put(self, request, grade_id):

        school = request.school

        audit_logger.info(
            "grade_update_started",
            performed_by=str(request.user.id),
            grade_id=str(grade_id),
            school_id=str(school.id) if school else None,
        )

        grade = Grade.objects.filter(id=grade_id, school=school).first()

        if grade is None:
            audit_logger.warning(
                "grade_update_failed",
                performed_by=str(request.user.id),
                grade_id=str(grade_id),
                school_id=str(school.id) if school else None,
                reason="grade_not_found",
            )
            return CustomResponse.errorResponse(description="Grade not found.")

        old_name = grade.name
        old_display_order = grade.display_order
        old_status = grade.status

        grade.name = request.data.get("name", grade.name)
        grade.display_order = request.data.get("display_order", grade.display_order)
        grade.status = request.data.get("status", grade.status)
        grade.save()

        audit_logger.info(
            "grade_updated",
            performed_by=str(request.user.id),
            grade_id=str(grade.id),
            school_id=str(school.id),
            academic_year_id=str(grade.academic_year_id),
            old_name=old_name,
            new_name=grade.name,
            old_display_order=old_display_order,
            new_display_order=grade.display_order,
            old_status=old_status,
            new_status=grade.status,
        )

        return CustomResponse.successResponse(description="Grade updated successfully.")


class CreateSectionAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "section.create"

    def post(self, request):

        school = request.school

        audit_logger.info(
            "section_create_started",
            performed_by=str(request.user.id),
            school_id=str(school.id) if school else None,
        )

        grade = Grade.objects.filter(id=request.data.get("grade_id"), school=school).first()

        if grade is None:
            audit_logger.warning(
                "section_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id) if school else None,
                grade_id=request.data.get("grade_id"),
                reason="grade_not_found",
            )
            return CustomResponse.errorResponse(description="Grade not found.")

        branch = None
        branch_id = request.data.get("branch_id")

        if branch_id:
            branch = Branch.objects.filter(id=branch_id, school=school).first()

            if branch is None:
                audit_logger.warning(
                    "section_create_failed",
                    performed_by=str(request.user.id),
                    school_id=str(school.id),
                    grade_id=str(grade.id),
                    branch_id=branch_id,
                    reason="branch_not_found",
                )
                return CustomResponse.errorResponse(description="Branch not found.")

        section_name = request.data.get("name")

        if Section.objects.filter(grade=grade, branch=branch, name=section_name).exists():
            audit_logger.warning(
                "section_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                grade_id=str(grade.id),
                branch_id=str(branch.id) if branch else None,
                section_name=section_name,
                reason="section_already_exists",
            )
            return CustomResponse.errorResponse(description="Section already exists.")

        class_teacher = None
        class_teacher_id = request.data.get("class_teacher_id")

        if class_teacher_id:
            class_teacher = Staff.objects.filter(
                id=class_teacher_id,
                school=school,
                branch=branch,
                staff_type=Staff.StaffType.TEACHER,
            ).first()

            if class_teacher is None:
                audit_logger.warning(
                    "section_create_failed",
                    performed_by=str(request.user.id),
                    school_id=str(school.id),
                    grade_id=str(grade.id),
                    branch_id=str(branch.id) if branch else None,
                    class_teacher_id=class_teacher_id,
                    reason="invalid_class_teacher",
                )
                return CustomResponse.errorResponse(description="Invalid class teacher.")

        section = Section.objects.create(
            grade=grade,
            branch=branch,
            name=section_name,
            class_teacher=class_teacher,
            capacity=request.data.get("capacity", 40),
            status=request.data.get("status", Section.Status.ACTIVE),
        )

        audit_logger.info(
            "section_created",
            performed_by=str(request.user.id),
            section_id=str(section.id),
            school_id=str(school.id),
            grade_id=str(grade.id),
            branch_id=str(branch.id) if branch else None,
            class_teacher_id=str(class_teacher.id) if class_teacher else None,
            section_name=section.name,
            capacity=section.capacity,
            status=section.status,
        )

        return CustomResponse.successResponse(
            description="Section created successfully.",
            data={"id": str(section.id)},
        )


class SectionListAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "section.view"

    def get(self, request):

        school = request.school
        branch_id = request.query_params.get("branch_id")

        application_logger.info(
            "sections_fetch_started",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            branch_id=branch_id,
        )

        sections = Section.objects.select_related("grade", "grade__school", "branch").filter(grade__school=school)

        if branch_id:
            sections = sections.filter(branch_id=branch_id)

        data = [
            {
                "id": str(section.id),
                "school": section.grade.school.name,
                "branch": {
                    "id": str(section.branch.id),
                    "name": section.branch.name,
                } if section.branch else None,
                "grade": section.grade.name,
                "name": section.name,
                "capacity": section.capacity,
                "status": section.status,
            }
            for section in sections
        ]

        application_logger.info(
            "sections_fetched",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            branch_id=branch_id,
            total_count=len(data),
        )

        return CustomResponse.successResponse(data=data)


class UpdateSectionAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "section.update"

    def put(self, request, section_id):

        school = request.school

        audit_logger.info(
            "section_update_started",
            performed_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            section_id=str(section_id),
        )

        if school is None:
            audit_logger.warning(
                "section_update_failed",
                performed_by=str(request.user.id),
                section_id=str(section_id),
                reason="school_not_found",
            )
            return CustomResponse.errorResponse(description="School not found.")

        section = Section.objects.select_related("grade", "branch", "class_teacher").filter(
            id=section_id,
            grade__school=school,
        ).first()

        if section is None:
            audit_logger.warning(
                "section_update_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                section_id=str(section_id),
                reason="section_not_found",
            )
            return CustomResponse.errorResponse(description="Section not found.")

        old_name = section.name
        old_branch_id = str(section.branch_id) if section.branch_id else None
        old_capacity = section.capacity
        old_status = section.status
        old_class_teacher_id = str(section.class_teacher_id) if section.class_teacher_id else None

        branch = section.branch

        if "branch_id" in request.data:
            branch_id = request.data.get("branch_id")

            if branch_id in [None, ""]:
                branch = None
            else:
                branch = Branch.objects.filter(id=branch_id, school=school).first()

                if branch is None:
                    audit_logger.warning(
                        "section_update_failed",
                        performed_by=str(request.user.id),
                        school_id=str(school.id),
                        section_id=str(section.id),
                        branch_id=branch_id,
                        reason="branch_not_found",
                    )
                    return CustomResponse.errorResponse(description="Branch not found.")

        name = request.data.get("name", section.name)

        if Section.objects.filter(grade=section.grade, branch=branch, name=name).exclude(id=section.id).exists():
            audit_logger.warning(
                "section_update_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                section_id=str(section.id),
                grade_id=str(section.grade_id),
                branch_id=str(branch.id) if branch else None,
                section_name=name,
                reason="section_already_exists",
            )
            return CustomResponse.errorResponse(description="Section already exists.")

        capacity = request.data.get("capacity", section.capacity)

        try:
            capacity = int(capacity)
        except (TypeError, ValueError):
            audit_logger.warning(
                "section_update_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                section_id=str(section.id),
                reason="invalid_capacity",
            )
            return CustomResponse.errorResponse(description="Invalid capacity.")

        if capacity <= 0:
            audit_logger.warning(
                "section_update_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                section_id=str(section.id),
                capacity=capacity,
                reason="capacity_must_be_greater_than_zero",
            )
            return CustomResponse.errorResponse(description="Capacity must be greater than zero.")

        status = request.data.get("status", section.status)

        if status not in Section.Status.values:
            audit_logger.warning(
                "section_update_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                section_id=str(section.id),
                status=status,
                reason="invalid_status",
            )
            return CustomResponse.errorResponse(description="Invalid status.")

        if "class_teacher_id" in request.data:
            class_teacher_id = request.data.get("class_teacher_id")

            if class_teacher_id:
                class_teacher = Staff.objects.filter(
                    id=class_teacher_id,
                    school=school,
                    branch=branch,
                    staff_type=Staff.StaffType.TEACHER,
                ).first()

                if class_teacher is None:
                    audit_logger.warning(
                        "section_update_failed",
                        performed_by=str(request.user.id),
                        school_id=str(school.id),
                        section_id=str(section.id),
                        class_teacher_id=class_teacher_id,
                        branch_id=str(branch.id) if branch else None,
                        reason="invalid_class_teacher",
                    )
                    return CustomResponse.errorResponse(description="Invalid class teacher.")

                section.class_teacher = class_teacher
            else:
                section.class_teacher = None

        elif section.class_teacher:
            teacher_branch_id = section.class_teacher.branch_id
            new_branch_id = branch.id if branch else None

            if teacher_branch_id != new_branch_id:
                section.class_teacher = None

        section.branch = branch
        section.name = name
        section.capacity = capacity
        section.status = status
        section.save()

        audit_logger.info(
            "section_updated",
            performed_by=str(request.user.id),
            school_id=str(school.id),
            section_id=str(section.id),
            grade_id=str(section.grade_id),
            old_name=old_name,
            new_name=section.name,
            old_branch_id=old_branch_id,
            new_branch_id=str(section.branch_id) if section.branch_id else None,
            old_capacity=old_capacity,
            new_capacity=section.capacity,
            old_status=old_status,
            new_status=section.status,
            old_class_teacher_id=old_class_teacher_id,
            new_class_teacher_id=str(section.class_teacher_id) if section.class_teacher_id else None,
        )

        return CustomResponse.successResponse(description="Section updated successfully.")


class CreateStudentAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "student.create"

    def post(self, request):

        school = request.school

        audit_logger.info(
            "student_create_started",
            performed_by=str(request.user.id),
            school_id=str(school.id) if school else None,
        )

        if school is None:
            audit_logger.warning(
                "student_create_failed",
                performed_by=str(request.user.id),
                reason="school_not_found",
            )
            return CustomResponse.errorResponse(description="School not found.")

        required_fields = [
            "academic_year_id",
            "grade_id",
            "section_id",
            "admission_number",
            "roll_number",
            "name",
            "gender",
            "date_of_birth",
            "admission_date",
        ]

        for field in required_fields:
            if request.data.get(field) in [None, ""]:
                audit_logger.warning(
                    "student_create_failed",
                    performed_by=str(request.user.id),
                    school_id=str(school.id),
                    missing_field=field,
                    reason="required_field_missing",
                )
                return CustomResponse.errorResponse(description=f"{field} is required.")

        branch = None
        branch_id = request.data.get("branch_id")

        if branch_id:
            branch = Branch.objects.filter(id=branch_id, school=school).first()

            if branch is None:
                audit_logger.warning(
                    "student_create_failed",
                    performed_by=str(request.user.id),
                    school_id=str(school.id),
                    branch_id=branch_id,
                    reason="branch_not_found",
                )
                return CustomResponse.errorResponse(description="Branch not found.")

        academic_year = AcademicYear.objects.filter(
            id=request.data.get("academic_year_id"),
            school=school,
        ).first()

        if academic_year is None:
            audit_logger.warning(
                "student_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                academic_year_id=request.data.get("academic_year_id"),
                reason="academic_year_not_found",
            )
            return CustomResponse.errorResponse(description="Academic year not found.")

        grade = Grade.objects.filter(
            id=request.data.get("grade_id"),
            school=school,
        ).first()

        if grade is None:
            audit_logger.warning(
                "student_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                grade_id=request.data.get("grade_id"),
                reason="grade_not_found",
            )
            return CustomResponse.errorResponse(description="Grade not found.")

        section = Section.objects.filter(
            id=request.data.get("section_id"),
            grade=grade,
            branch=branch,
        ).first()

        if section is None:
            audit_logger.warning(
                "student_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                branch_id=str(branch.id) if branch else None,
                grade_id=str(grade.id),
                section_id=request.data.get("section_id"),
                reason="section_not_found",
            )
            return CustomResponse.errorResponse(description="Section not found.")

        gender = request.data.get("gender")

        if gender not in Student.Gender.values:
            audit_logger.warning(
                "student_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                gender=gender,
                reason="invalid_gender",
            )
            return CustomResponse.errorResponse(description="Invalid gender.")

        board = request.data.get("board", Student.Board.STATE)

        if board not in Student.Board.values:
            audit_logger.warning(
                "student_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                board=board,
                reason="invalid_board",
            )
            return CustomResponse.errorResponse(description="Invalid board.")

        hostel_type = request.data.get("hostel_type", Student.HostelType.DAY_SCHOLAR)

        if hostel_type not in Student.HostelType.values:
            audit_logger.warning(
                "student_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                hostel_type=hostel_type,
                reason="invalid_hostel_type",
            )
            return CustomResponse.errorResponse(description="Invalid hostel type.")

        enrollment_type = request.data.get("enrollment_type", Student.EnrollmentType.NEW)

        if enrollment_type not in Student.EnrollmentType.values:
            audit_logger.warning(
                "student_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                enrollment_type=enrollment_type,
                reason="invalid_enrollment_type",
            )
            return CustomResponse.errorResponse(description="Invalid enrollment type.")

        admission_number = request.data.get("admission_number").strip()

        if Student.objects.filter(school=school, admission_number=admission_number).exists():
            audit_logger.warning(
                "student_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                admission_number=admission_number,
                reason="admission_number_already_exists",
            )
            return CustomResponse.errorResponse(description="Admission number already exists.")

        roll_number = request.data.get("roll_number")

        if Student.objects.filter(section=section, roll_number=roll_number).exists():
            audit_logger.warning(
                "student_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                section_id=str(section.id),
                roll_number=roll_number,
                reason="roll_number_already_exists",
            )
            return CustomResponse.errorResponse(description="Roll number already exists in this section.")

        valid_blood_groups = ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]
        blood_group = request.data.get("blood_group")

        if blood_group and blood_group not in valid_blood_groups:
            audit_logger.warning(
                "student_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                blood_group=blood_group,
                reason="invalid_blood_group",
            )
            return CustomResponse.errorResponse(description="Invalid blood group.")

        try:
            with transaction.atomic():

                fee_concession_id = request.data.get("fee_concession_id")
                concession = None

                if fee_concession_id:
                    concession = FeeConcession.objects.filter(
                        id=fee_concession_id,
                        school=school,
                    ).first()

                    if concession is None:
                        audit_logger.warning(
                            "student_create_failed",
                            performed_by=str(request.user.id),
                            school_id=str(school.id),
                            fee_concession_id=fee_concession_id,
                            reason="fee_concession_not_found",
                        )
                        return CustomResponse.errorResponse(description="Fee concession not found.")

                student = Student.objects.create(
                    school=school,
                    branch=branch,
                    board=board,
                    academic_year=academic_year,
                    grade=grade,
                    section=section,
                    admission_number=admission_number,
                    roll_number=roll_number,
                    name=request.data.get("name").strip(),
                    gender=gender,
                    date_of_birth=request.data.get("date_of_birth"),
                    admission_date=request.data.get("admission_date"),
                    enrollment_type=enrollment_type,
                    status=request.data.get("status", Student.Status.ACTIVE),
                    place_of_birth=request.data.get("place_of_birth"),
                    blood_group=blood_group,
                    photo_url=request.data.get("photo_url"),
                    nationality=request.data.get("nationality", "Indian"),
                    mother_tongue=request.data.get("mother_tongue"),
                    aadhaar_number=request.data.get("aadhaar_number"),
                    religion=request.data.get("religion"),
                    caste=request.data.get("caste"),
                    sub_caste=request.data.get("sub_caste"),
                    student_category=request.data.get("student_category"),
                    identification_marks=request.data.get("identification_marks"),
                    email=request.data.get("email"),
                    address=request.data.get("address"),
                    emergency_contact_name=request.data.get("emergency_contact_name"),
                    emergency_contact_mobile=request.data.get("emergency_contact_mobile"),
                    father_name=request.data.get("father_name"),
                    father_mobile=request.data.get("father_mobile"),
                    father_occupation=request.data.get("father_occupation"),
                    mother_name=request.data.get("mother_name"),
                    mother_mobile=request.data.get("mother_mobile"),
                    mother_occupation=request.data.get("mother_occupation"),
                    guardian_name=request.data.get("guardian_name"),
                    guardian_mobile=request.data.get("guardian_mobile"),
                    guardian_occupation=request.data.get("guardian_occupation"),
                    previous_school_name=request.data.get("previous_school_name"),
                    previous_school_tc_number=request.data.get("previous_school_tc_number"),
                    previous_exam_percentage=request.data.get("previous_exam_percentage"),
                    transport_required=request.data.get("transport_required", False),
                    pickup_point=request.data.get("pickup_point"),
                    hostel_type=hostel_type,
                )

                fee_template = FeeTemplate.objects.filter(
                    school=school,
                    academic_year=academic_year,
                    grade=grade,
                ).first()

                if fee_template is None:
                    raise Exception(f"Fee template not configured for grade '{grade.name}'.")

                StudentFeeAssignment.objects.create(
                    student=student,
                    fee_template=fee_template,
                    concession=concession,
                    assigned_by=request.user,
                )

                generate_student_fees(
                    student=student,
                    fee_template=fee_template,
                )

        except Exception as e:
            audit_logger.exception(
                "student_create_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                branch_id=str(branch.id) if branch else None,
                academic_year_id=str(academic_year.id),
                grade_id=str(grade.id),
                section_id=str(section.id),
                admission_number=admission_number,
                reason="exception",
                error=str(e),
            )
            return CustomResponse.errorResponse(description=str(e))

        audit_logger.info(
            "student_created",
            performed_by=str(request.user.id),
            student_id=str(student.id),
            school_id=str(school.id),
            branch_id=str(branch.id) if branch else None,
            academic_year_id=str(academic_year.id),
            grade_id=str(grade.id),
            section_id=str(section.id),
            admission_number=student.admission_number,
            roll_number=student.roll_number,
            student_name=student.name,
            fee_template_id=str(fee_template.id),
            concession_id=str(concession.id) if concession else None,
        )

        return CustomResponse.successResponse(
            description="Student created successfully.",
            data={
                "id": str(student.id),
                "name": student.name,
                "admission_number": student.admission_number,
            },
        )


class BulkUploadStudentAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "student.bulk_upload"
    parser_classes = [MultiPartParser]

    def post(self, request):

        school = request.school

        audit_logger.info(
            "student_bulk_upload_started",
            performed_by=str(request.user.id),
            school_id=str(school.id) if school else None,
        )

        if school is None:
            audit_logger.warning(
                "student_bulk_upload_failed",
                performed_by=str(request.user.id),
                reason="school_not_found",
            )
            return CustomResponse.errorResponse(description="School not found.")

        file = request.FILES.get("file")

        if file is None:
            audit_logger.warning(
                "student_bulk_upload_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                reason="excel_file_missing",
            )
            return CustomResponse.errorResponse(description="Excel file is required.")

        audit_logger.info(
            "student_bulk_upload_file_received",
            performed_by=str(request.user.id),
            school_id=str(school.id),
            file_name=file.name,
            file_size=file.size,
        )

        try:
            workbook = load_workbook(file)
            sheet = workbook.active

        except Exception as e:
            audit_logger.exception(
                "student_bulk_upload_failed",
                performed_by=str(request.user.id),
                school_id=str(school.id),
                file_name=file.name,
                reason="invalid_excel_file",
                error=str(e),
            )
            return CustomResponse.errorResponse(description="Invalid Excel file.")

        headers = [cell.value for cell in sheet[1]]

        required_headers = [
            "Academic Year",
            "Grade",
            "Section",
            "Admission Number",
            "Roll Number",
            "Name",
            "Gender",
            "Date Of Birth",
            "Admission Date",
            "Fee Concession",
        ]

        for header in required_headers:
            if header not in headers:
                audit_logger.warning(
                    "student_bulk_upload_failed",
                    performed_by=str(request.user.id),
                    school_id=str(school.id),
                    file_name=file.name,
                    missing_header=header,
                    reason="required_column_missing",
                )
                return CustomResponse.errorResponse(description=f"{header} column is missing.")

        success_count = 0
        failed_count = 0
        errors = []

        valid_blood_groups = ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

            try:
                data = dict(zip(headers, row))

                academic_year = AcademicYear.objects.filter(
                    school=school,
                    name=data.get("Academic Year"),
                ).first()

                if academic_year is None:
                    raise Exception("Academic year not found.")

                grade = Grade.objects.filter(
                    school=school,
                    name=data.get("Grade"),
                ).first()

                if grade is None:
                    raise Exception("Grade not found.")

                branch = None
                branch_name = str(data.get("Branch") or "").strip()

                if branch_name:
                    branch = Branch.objects.filter(
                        school=school,
                        name__iexact=branch_name,
                    ).first()

                    if branch is None:
                        raise Exception(f"Branch '{branch_name}' not found.")

                section = Section.objects.filter(
                    grade=grade,
                    branch=branch,
                    name__iexact=str(data.get("Section")).strip(),
                ).first()

                if section is None:
                    raise Exception("Section not found.")

                gender = str(data.get("Gender")).upper()

                if gender not in Student.Gender.values:
                    raise Exception("Invalid gender.")

                board = str(data.get("Board", Student.Board.STATE)).upper()

                if board not in Student.Board.values:
                    raise Exception("Invalid board.")

                blood_group = data.get("Blood Group")

                if blood_group and blood_group not in valid_blood_groups:
                    raise Exception("Invalid blood group.")

                admission_number = str(data.get("Admission Number")).strip()
                roll_number = data.get("Roll Number")

                if Student.objects.filter(
                    school=school,
                    admission_number=admission_number,
                ).exists():
                    raise Exception("Admission number already exists.")

                if Student.objects.filter(
                    section=section,
                    roll_number=roll_number,
                ).exists():
                    raise Exception("Roll number already exists.")

                enrollment_type = str(
                    data.get("Enrollment Type", Student.EnrollmentType.NEW)
                ).upper()

                if enrollment_type not in Student.EnrollmentType.values:
                    raise Exception("Invalid enrollment type.")

                hostel_type = str(
                    data.get("Hostel Type", Student.HostelType.DAY_SCHOLAR)
                ).upper()

                if hostel_type not in Student.HostelType.values:
                    raise Exception("Invalid hostel type.")

                with transaction.atomic():

                    student = Student.objects.create(
                        school=school,
                        branch=branch,
                        board=board,
                        academic_year=academic_year,
                        grade=grade,
                        section=section,
                        admission_number=admission_number,
                        roll_number=roll_number,
                        name=str(data.get("Name")).strip(),
                        gender=gender,
                        date_of_birth=data.get("Date Of Birth"),
                        admission_date=data.get("Admission Date"),
                        enrollment_type=enrollment_type,
                        place_of_birth=data.get("Place Of Birth"),
                        blood_group=blood_group,
                        nationality=data.get("Nationality", "Indian"),
                        mother_tongue=data.get("Mother Tongue"),
                        aadhaar_number=data.get("Aadhaar Number"),
                        religion=data.get("Religion"),
                        caste=data.get("Caste"),
                        sub_caste=data.get("Sub Caste"),
                        student_category=data.get("Student Category"),
                        email=data.get("Email"),
                        address=data.get("Address"),
                        father_name=data.get("Father Name"),
                        father_mobile=data.get("Father Mobile"),
                        father_occupation=data.get("Father Occupation"),
                        mother_name=data.get("Mother Name"),
                        mother_mobile=data.get("Mother Mobile"),
                        mother_occupation=data.get("Mother Occupation"),
                        guardian_name=data.get("Guardian Name"),
                        guardian_mobile=data.get("Guardian Mobile"),
                        guardian_occupation=data.get("Guardian Occupation"),
                        previous_school_name=data.get("Previous School"),
                        previous_exam_percentage=data.get("Previous Exam Percentage"),
                        transport_required=bool(data.get("Transport Required", False)),
                        pickup_point=data.get("Pickup Point"),
                        hostel_type=hostel_type,
                        status=Student.Status.ACTIVE,
                    )

                    fee_concession_name = str(data.get("Fee Concession", "")).strip()
                    concession = None

                    if fee_concession_name:
                        concession = FeeConcession.objects.filter(
                            school=school,
                            name__iexact=fee_concession_name,
                        ).first()

                        if concession is None:
                            raise Exception(f"Fee concession '{fee_concession_name}' not found.")

                    fee_template = FeeTemplate.objects.filter(
                        school=school,
                        academic_year=academic_year,
                        grade=grade,
                        is_active=True,
                    ).first()

                    if fee_template is None:
                        raise Exception(f"Fee template not configured for grade '{grade.name}'.")

                    StudentFeeAssignment.objects.get_or_create(
                        student=student,
                        fee_template=fee_template,
                        defaults={
                            "concession": concession,
                            "assigned_by": request.user,
                        },
                    )

                    generate_student_fees(
                        student=student,
                        fee_template=fee_template,
                    )

                success_count += 1

            except Exception as e:

                failed_count += 1

                errors.append({
                    "row": row_number,
                    "message": str(e),
                })

                audit_logger.warning(
                    "student_bulk_upload_row_failed",
                    performed_by=str(request.user.id),
                    school_id=str(school.id),
                    file_name=file.name,
                    row_number=row_number,
                    admission_number=str(data.get("Admission Number")) if "data" in locals() else None,
                    student_name=str(data.get("Name")) if "data" in locals() else None,
                    branch_name=str(data.get("Branch")) if "data" in locals() else None,
                    grade_name=str(data.get("Grade")) if "data" in locals() else None,
                    section_name=str(data.get("Section")) if "data" in locals() else None,
                    error=str(e),
                )

        audit_logger.info(
            "student_bulk_upload_completed",
            performed_by=str(request.user.id),
            school_id=str(school.id),
            file_name=file.name,
            total_records=success_count + failed_count,
            success_count=success_count,
            failed_count=failed_count,
        )

        return CustomResponse.successResponse(
            description="Student bulk upload completed.",
            data={
                "total_records": success_count + failed_count,
                "success_count": success_count,
                "failed_count": failed_count,
                "errors": errors,
            },
        )


class DownloadStudentTemplateAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "student.bulk_upload"

    def get(self, request):

        school = request.school

        application_logger.info(
            "student_template_download_started",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
        )

        try:

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Students"

            headers = [
                "Academic Year",
                "Branch",
                "Grade",
                "Section",
                "Board",
                "Admission Number",
                "Roll Number",
                "Name",
                "Gender",
                "Date Of Birth",
                "Admission Date",
                "Enrollment Type",
                "Place Of Birth",
                "Blood Group",
                "Nationality",
                "Mother Tongue",
                "Aadhaar Number",
                "Religion",
                "Caste",
                "Sub Caste",
                "Student Category",
                "Identification Marks",
                "Email",
                "Address",
                "Emergency Contact Name",
                "Emergency Contact Mobile",
                "Father Name",
                "Father Mobile",
                "Father Occupation",
                "Mother Name",
                "Mother Mobile",
                "Mother Occupation",
                "Guardian Name",
                "Guardian Mobile",
                "Guardian Occupation",
                "Previous School",
                "Previous School TC Number",
                "Previous Exam Percentage",
                "Transport Required",
                "Pickup Point",
                "Hostel Type",
                "Fee Concession",
            ]

            for column_number, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=column_number, value=header)

            sample_row = [
                "2025-2026",
                "Main Branch",
                "Grade 1",
                "A",
                "CBSE",
                "ADM001",
                1,
                "Rahul Kumar",
                "MALE",
                "2018-05-10",
                "2025-06-01",
                "NEW",
                "Hyderabad",
                "O+",
                "Indian",
                "Telugu",
                "123456789012",
                "Hindu",
                "OC",
                "Kamma",
                "General",
                "Mole on chin",
                "rahul@gmail.com",
                "Hyderabad",
                "Ramesh",
                "9876543210",
                "Ramesh",
                "9876543210",
                "Engineer",
                "Sita",
                "9876543211",
                "Teacher",
                "Ramesh",
                "9876543210",
                "Engineer",
                "ABC School",
                "TC123",
                92.5,
                True,
                "Miyapur",
                "DAY_SCHOLAR",
            ]

            for column_number, value in enumerate(sample_row, start=1):
                sheet.cell(row=2, column=column_number, value=value)

            excel_file = BytesIO()
            workbook.save(excel_file)
            excel_file.seek(0)

            file_name = "templates/student_template.xlsx"

            file_path = default_storage.save(
                file_name,
                ContentFile(excel_file.getvalue()),
            )

            file_url = settings.MEDIA_URL + file_path

        except Exception as e:

            application_logger.exception(
                "student_template_download_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id) if school else None,
                reason="template_generation_failed",
                error=str(e),
            )

            return CustomResponse.errorResponse(
                description=str(e),
            )

        application_logger.info(
            "student_template_generated",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            file_path=file_path,
            file_url=file_url,
            total_columns=len(headers),
        )

        return CustomResponse.successResponse(
            description="Student template generated successfully.",
            data={
                "file_url": file_url,
                "file_path": file_path,
            },
        )

class StudentListAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "student.view"
    pagination_class = CustomPageNumberPagination

    def get(self, request):

        school = request.school

        academic_year_id = request.query_params.get("academic_year_id")
        branch_id = request.query_params.get("branch_id")
        grade_id = request.query_params.get("grade_id")
        section_id = request.query_params.get("section_id")
        board = request.query_params.get("board")
        hostel_type = request.query_params.get("hostel_type")
        student_status = request.query_params.get("status")
        search = request.query_params.get("search")

        application_logger.info(
            "student_list_requested",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            academic_year_id=academic_year_id,
            branch_id=branch_id,
            grade_id=grade_id,
            section_id=section_id,
            board=board,
            hostel_type=hostel_type,
            status=student_status,
            search=search,
        )

        try:

            students = Student.objects.select_related(
                "school",
                "branch",
                "academic_year",
                "grade",
                "section",
            ).filter(
                school=school,
            )

            if academic_year_id:
                students = students.filter(academic_year_id=academic_year_id)

            if branch_id:
                students = students.filter(branch_id=branch_id)

            if grade_id:
                students = students.filter(grade_id=grade_id)

            if section_id:
                students = students.filter(section_id=section_id)

            if board:
                students = students.filter(board=board)

            if hostel_type:
                students = students.filter(hostel_type=hostel_type)

            if student_status:
                students = students.filter(status=student_status)

            if search:
                students = students.filter(
                    Q(name__icontains=search)
                    | Q(admission_number__icontains=search)
                    | Q(father_mobile__icontains=search)
                    | Q(mother_mobile__icontains=search)
                )

            students = students.order_by("roll_number")

            total = students.count()

            paginator = self.pagination_class()
            page = paginator.paginate_queryset(students, request)

            data = []

            for student in page:

                data.append({
                    "id": str(student.id),
                    "admission_number": student.admission_number,
                    "roll_number": student.roll_number,
                    "name": student.name,
                    "board": student.board,
                    "gender": student.gender,
                    "date_of_birth": student.date_of_birth,
                    "blood_group": student.blood_group,
                    "status": student.status,
                    "school": {
                        "id": str(student.school.id),
                        "name": student.school.name,
                    },
                    "academic_year": {
                        "id": str(student.academic_year.id),
                        "name": student.academic_year.name,
                    },
                    "branch": {
                        "id": str(student.branch.id),
                        "name": student.branch.name,
                    } if student.branch else None,
                    "grade": {
                        "id": str(student.grade.id),
                        "name": student.grade.name,
                    },
                    "section": {
                        "id": str(student.section.id),
                        "name": student.section.name,
                    },
                    "father_name": student.father_name,
                    "father_mobile": student.father_mobile,
                    "mother_name": student.mother_name,
                    "mother_mobile": student.mother_mobile,
                    "guardian_name": student.guardian_name,
                    "guardian_mobile": student.guardian_mobile,
                    "hostel_type": student.hostel_type,
                    "transport_required": student.transport_required,
                })

        except Exception as e:

            application_logger.exception(
                "student_list_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id) if school else None,
                academic_year_id=academic_year_id,
                branch_id=branch_id,
                grade_id=grade_id,
                section_id=section_id,
                reason="student_list_fetch_failed",
                error=str(e),
            )

            return CustomResponse.errorResponse(
                description=str(e),
            )

        application_logger.info(
            "student_list_fetched",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            total=total,
            returned_count=len(data),
            academic_year_id=academic_year_id,
            branch_id=branch_id,
            grade_id=grade_id,
            section_id=section_id,
        )

        return CustomResponse.successResponse(
            data=data,
            total=total,
        )



class CreateStudentDocumentAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "student_document.create"

    def post(self, request):

        school = request.school
        student_id = request.data.get("student_id")
        academic_year_id = request.data.get("academic_year_id")
        document_type = request.data.get("document_type")

        application_logger.info(
            "student_document_create_requested",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            student_id=student_id,
            academic_year_id=academic_year_id,
            document_type=document_type,
        )

        student = Student.objects.filter(
            id=student_id,
            school=school,
        ).first()

        if student is None:

            application_logger.warning(
                "student_document_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id) if school else None,
                student_id=student_id,
                reason="student_not_found",
            )

            return CustomResponse.errorResponse(
                description="Student not found.",
            )

        academic_year = None

        if academic_year_id:

            academic_year = AcademicYear.objects.filter(
                id=academic_year_id,
                school=school,
            ).first()

            if academic_year is None:

                application_logger.warning(
                    "student_document_create_failed",
                    requested_by=str(request.user.id),
                    school_id=str(school.id) if school else None,
                    student_id=student_id,
                    academic_year_id=academic_year_id,
                    reason="academic_year_not_found",
                )

                return CustomResponse.errorResponse(
                    description="Academic year not found.",
                )

        if document_type not in StudentDocument.DocumentType.values:

            application_logger.warning(
                "student_document_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id) if school else None,
                student_id=student_id,
                document_type=document_type,
                reason="invalid_document_type",
            )

            return CustomResponse.errorResponse(
                description="Invalid document type.",
            )

        try:

            document = StudentDocument.objects.create(
                student=student,
                academic_year=academic_year,
                document_type=document_type,
                title=request.data.get("title"),
                file_url=request.data.get("file_url"),
                remarks=request.data.get("remarks"),
                status=request.data.get(
                    "status",
                    StudentDocument.Status.ACTIVE,
                ),
            )

        except Exception as e:

            application_logger.exception(
                "student_document_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id) if school else None,
                student_id=student_id,
                academic_year_id=academic_year_id,
                document_type=document_type,
                reason="student_document_creation_failed",
                error=str(e),
            )

            return CustomResponse.errorResponse(
                description=str(e),
            )

        application_logger.info(
            "student_document_created",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            student_id=str(student.id),
            document_id=str(document.id),
            document_type=document.document_type,
        )

        return CustomResponse.successResponse(
            description="Student document created successfully.",
            data={
                "id": str(document.id),
            },
        )


class StudentDocumentListAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "student_document.view"

    def get(self, request):

        school = request.school

        student_id = request.query_params.get("student_id")
        academic_year_id = request.query_params.get("academic_year_id")
        document_type = request.query_params.get("document_type")
        document_status = request.query_params.get("status")

        application_logger.info(
            "student_document_list_requested",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            student_id=student_id,
            academic_year_id=academic_year_id,
            document_type=document_type,
            status=document_status,
        )

        try:

            documents = StudentDocument.objects.select_related(
                "student",
                "academic_year",
            ).filter(
                student__school=school,
            )

            if student_id:
                documents = documents.filter(student_id=student_id)

            if academic_year_id:
                documents = documents.filter(
                    academic_year_id=academic_year_id
                )

            if document_type:
                documents = documents.filter(
                    document_type=document_type
                )

            if document_status:
                documents = documents.filter(
                    status=document_status
                )

            data = []

            for document in documents:

                data.append({
                    "id": str(document.id),
                    "student": {
                        "id": str(document.student.id),
                        "name": document.student.name,
                    },
                    "academic_year": {
                        "id": str(document.academic_year.id),
                        "name": document.academic_year.name,
                    } if document.academic_year else None,
                    "document_type": document.document_type,
                    "title": document.title,
                    "file_url": document.file_url,
                    "remarks": document.remarks,
                    "status": document.status,
                })

        except Exception as e:

            application_logger.exception(
                "student_document_list_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id) if school else None,
                student_id=student_id,
                academic_year_id=academic_year_id,
                document_type=document_type,
                reason="student_document_list_fetch_failed",
                error=str(e),
            )

            return CustomResponse.errorResponse(
                description=str(e),
            )

        application_logger.info(
            "student_document_list_fetched",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            student_id=student_id,
            academic_year_id=academic_year_id,
            returned_count=len(data),
        )

        return CustomResponse.successResponse(
            data=data,
        )


class UpdateStudentDocumentAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "student_document.update"

    def put(self, request, document_id):

        school = request.school

        application_logger.info(
            "student_document_update_requested",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            document_id=str(document_id),
        )

        document = StudentDocument.objects.select_related(
            "student",
            "academic_year",
        ).filter(
            id=document_id,
            student__school=school,
        ).first()

        if document is None:

            application_logger.warning(
                "student_document_update_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id) if school else None,
                document_id=str(document_id),
                reason="student_document_not_found",
            )

            return CustomResponse.errorResponse(
                description="Student document not found.",
            )

        academic_year_id = request.data.get("academic_year_id")
        document_type = request.data.get("document_type")

        if academic_year_id:

            academic_year = AcademicYear.objects.filter(
                id=academic_year_id,
                school=school,
            ).first()

            if academic_year is None:

                application_logger.warning(
                    "student_document_update_failed",
                    requested_by=str(request.user.id),
                    school_id=str(school.id) if school else None,
                    document_id=str(document.id),
                    academic_year_id=academic_year_id,
                    reason="academic_year_not_found",
                )

                return CustomResponse.errorResponse(
                    description="Academic year not found.",
                )

            document.academic_year = academic_year

        if document_type:

            if document_type not in StudentDocument.DocumentType.values:

                application_logger.warning(
                    "student_document_update_failed",
                    requested_by=str(request.user.id),
                    school_id=str(school.id) if school else None,
                    document_id=str(document.id),
                    document_type=document_type,
                    reason="invalid_document_type",
                )

                return CustomResponse.errorResponse(
                    description="Invalid document type.",
                )

            document.document_type = document_type

        try:

            document.title = request.data.get("title", document.title)
            document.file_url = request.data.get("file_url", document.file_url)
            document.remarks = request.data.get("remarks", document.remarks)
            document.status = request.data.get("status", document.status)

            document.save()

        except Exception as e:

            application_logger.exception(
                "student_document_update_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id) if school else None,
                document_id=str(document.id),
                reason="student_document_update_failed",
                error=str(e),
            )

            return CustomResponse.errorResponse(
                description=str(e),
            )

        application_logger.info(
            "student_document_updated",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            student_id=str(document.student_id),
            document_id=str(document.id),
            document_type=document.document_type,
            status=document.status,
        )

        return CustomResponse.successResponse(
            description="Student document updated successfully.",
        )


class CreateStaffAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "staff.create"

    def post(self, request):

        school = request.school
        employee_id = request.data.get("employee_id")
        staff_type = request.data.get("staff_type")
        mobile = str(request.data.get("mobile"))
        branch_id = request.data.get("branch_id")

        application_logger.info(
            "staff_create_requested",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            employee_id=employee_id,
            staff_type=staff_type,
            branch_id=branch_id,
        )

        if school is None:
            application_logger.warning("staff_create_failed", requested_by=str(request.user.id), reason="school_not_found")
            return CustomResponse.errorResponse(description="School not found.")

        required_fields = ["employee_id", "staff_type", "name", "gender", "mobile", "joining_date"]

        for field in required_fields:
            if request.data.get(field) in [None, ""]:
                application_logger.warning(
                    "staff_create_failed",
                    requested_by=str(request.user.id),
                    school_id=str(school.id),
                    field=field,
                    reason="required_field_missing",
                )
                return CustomResponse.errorResponse(description=f"{field} is required.")

        if staff_type not in Staff.StaffType.values:
            application_logger.warning(
                "staff_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_type=staff_type,
                reason="invalid_staff_type",
            )
            return CustomResponse.errorResponse(description="Invalid staff type.")

        gender = request.data.get("gender")

        if gender not in Staff.Gender.values:
            application_logger.warning(
                "staff_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                gender=gender,
                reason="invalid_gender",
            )
            return CustomResponse.errorResponse(description="Invalid gender.")

        if not mobile.isdigit() or len(mobile) != 10:
            application_logger.warning(
                "staff_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                reason="invalid_mobile_number",
            )
            return CustomResponse.errorResponse(description="Enter valid mobile number.")

        branch = None

        if branch_id:
            branch = Branch.objects.filter(id=branch_id, school=school).first()

            if branch is None:
                application_logger.warning(
                    "staff_create_failed",
                    requested_by=str(request.user.id),
                    school_id=str(school.id),
                    branch_id=branch_id,
                    reason="invalid_branch",
                )
                return CustomResponse.errorResponse(description="Invalid branch.")

        if UserMaster.objects.filter(username=mobile).exists():
            application_logger.warning(
                "staff_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                reason="user_already_exists",
            )
            return CustomResponse.errorResponse(description="User already exists.")

        if Staff.objects.filter(school=school, employee_id=employee_id).exists():
            application_logger.warning(
                "staff_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                employee_id=employee_id,
                reason="employee_id_already_exists",
            )
            return CustomResponse.errorResponse(description="Employee ID already exists.")

        if Staff.objects.filter(school=school, mobile=mobile).exists():
            application_logger.warning(
                "staff_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                reason="mobile_already_exists",
            )
            return CustomResponse.errorResponse(description="Mobile number already exists.")

        email = request.data.get("email")

        if email and Staff.objects.filter(school=school, email=email).exists():
            application_logger.warning(
                "staff_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                reason="email_already_exists",
            )
            return CustomResponse.errorResponse(description="Email already exists.")

        role = Roles.objects.filter(role_name=staff_type).first()

        if role is None:
            application_logger.warning(
                "staff_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_type=staff_type,
                reason="role_not_configured",
            )
            return CustomResponse.errorResponse(description=f"{staff_type.title()} role not configured.")

        try:

            with transaction.atomic():

                user = UserMaster.objects.create(
                    username=mobile,
                    first_name=request.data.get("name").strip(),
                    email=email,
                    mobile=mobile,
                    gender=gender,
                    date_of_birth=request.data.get("date_of_birth"),
                    profile_image=request.data.get("profile_image"),
                    is_active=True,
                )

                staff = Staff.objects.create(
                    user=user,
                    school=school,
                    branch=branch,
                    employee_id=employee_id.strip(),
                    staff_type=staff_type,
                    name=request.data.get("name").strip(),
                    gender=gender,
                    date_of_birth=request.data.get("date_of_birth"),
                    mobile=mobile,
                    email=email,
                    qualification=request.data.get("qualification"),
                    experience=request.data.get("experience", 0),
                    joining_date=request.data.get("joining_date"),
                    status=request.data.get("status", Staff.Status.ACTIVE),
                    profile_image=request.data.get("profile_image"),
                    address=request.data.get("address"),
                    emergency_contact_name=request.data.get("emergency_contact_name"),
                    emergency_contact_mobile=request.data.get("emergency_contact_mobile"),
                )

                UserRoles.objects.create(user=user, school=school, role=role)

        except Exception as e:
            application_logger.exception(
                "staff_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                employee_id=employee_id,
                branch_id=branch_id,
                reason="staff_creation_failed",
                error=str(e),
            )
            return CustomResponse.errorResponse(description=str(e))

        application_logger.info(
            "staff_created",
            requested_by=str(request.user.id),
            school_id=str(school.id),
            staff_id=str(staff.id),
            user_id=str(user.id),
            employee_id=staff.employee_id,
            staff_type=staff.staff_type,
            branch_id=str(staff.branch_id) if staff.branch_id else None,
            role=role.role_name,
        )

        return CustomResponse.successResponse(
            description="Staff created successfully.",
            data={
                "id": str(staff.id),
                "employee_id": staff.employee_id,
                "name": staff.name,
                "role": role.role_name,
            },
        )




class GetStaffAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "staff.view"

    def get(self, request):

        school = request.school
        branch_id = request.query_params.get("branch_id")

        application_logger.info(
            "staff_list_requested",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            branch_id=branch_id,
        )

        if school is None:

            application_logger.warning(
                "staff_list_failed",
                requested_by=str(request.user.id),
                reason="school_not_found",
            )

            return CustomResponse.errorResponse(
                description="School not found."
            )

        try:

            staffs = Staff.objects.select_related(
                "user",
                "branch",
            ).prefetch_related(
                Prefetch(
                    "user__user_roles",
                    queryset=UserRoles.objects.filter(
                        school=school,
                    ).select_related(
                        "role",
                    ),
                    to_attr="school_roles",
                )
            ).filter(
                school=school,
            )

            if branch_id:

                staffs = staffs.filter(
                    branch_id=branch_id,
                )

            staffs = staffs.order_by("name")

            data = []

            for staff in staffs:

                user_roles = staff.user.school_roles

                role = (
                    user_roles[0].role
                    if user_roles
                    else None
                )

                data.append({
                    "id": str(staff.id),
                    "employee_id": staff.employee_id,
                    "staff_type": staff.staff_type,
                    "role": role.role_name if role else None,
                    "name": staff.name,
                    "gender": staff.gender,
                    "date_of_birth": staff.date_of_birth,
                    "mobile": staff.mobile,
                    "email": staff.email,
                    "qualification": staff.qualification,
                    "experience": staff.experience,
                    "joining_date": staff.joining_date,
                    "status": staff.status,
                    "profile_image": staff.profile_image,
                    "address": staff.address,
                    "emergency_contact_name": staff.emergency_contact_name,
                    "emergency_contact_mobile": staff.emergency_contact_mobile,
                    "branch": {
                        "id": str(staff.branch.id),
                        "name": staff.branch.name,
                    } if staff.branch else None,
                    "created_at": staff.created_at,
                })

        except Exception as e:

            application_logger.exception(
                "staff_list_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                branch_id=branch_id,
                reason="staff_list_fetch_failed",
                error=str(e),
            )

            return CustomResponse.errorResponse(
                description=str(e)
            )

        application_logger.info(
            "staff_list_fetched",
            requested_by=str(request.user.id),
            school_id=str(school.id),
            branch_id=branch_id,
            returned_count=len(data),
        )

        return CustomResponse.successResponse(
            description="Staff fetched successfully.",
            data=data,
        )

class UpdateStaffAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "staff.update"

    def put(self, request, staff_id):

        school = request.school

        application_logger.info(
            "staff_update_requested",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            staff_id=str(staff_id),
        )

        if school is None:
            application_logger.warning(
                "staff_update_failed",
                requested_by=str(request.user.id),
                staff_id=str(staff_id),
                reason="school_not_found",
            )
            return CustomResponse.errorResponse(description="School not found.")

        staff = Staff.objects.select_related("user", "branch").filter(id=staff_id, school=school).first()

        if staff is None:
            application_logger.warning(
                "staff_update_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=str(staff_id),
                reason="staff_not_found",
            )
            return CustomResponse.errorResponse(description="Staff not found.")

        branch = staff.branch

        if "branch_id" in request.data:

            branch_id = request.data.get("branch_id")

            if branch_id in [None, ""]:
                branch = None
            else:
                branch = Branch.objects.filter(id=branch_id, school=school).first()

                if branch is None:
                    application_logger.warning(
                        "staff_update_failed",
                        requested_by=str(request.user.id),
                        school_id=str(school.id),
                        staff_id=str(staff.id),
                        branch_id=branch_id,
                        reason="invalid_branch",
                    )
                    return CustomResponse.errorResponse(description="Invalid branch.")

        employee_id = request.data.get("employee_id", staff.employee_id)

        if Staff.objects.filter(school=school, employee_id=employee_id).exclude(id=staff.id).exists():
            application_logger.warning(
                "staff_update_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=str(staff.id),
                employee_id=employee_id,
                reason="employee_id_already_exists",
            )
            return CustomResponse.errorResponse(description="Employee ID already exists.")

        mobile = str(request.data.get("mobile", staff.mobile))

        if not mobile.isdigit() or len(mobile) != 10:
            application_logger.warning(
                "staff_update_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=str(staff.id),
                reason="invalid_mobile_number",
            )
            return CustomResponse.errorResponse(description="Enter valid mobile number.")

        if Staff.objects.filter(school=school, mobile=mobile).exclude(id=staff.id).exists():
            application_logger.warning(
                "staff_update_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=str(staff.id),
                reason="mobile_already_exists",
            )
            return CustomResponse.errorResponse(description="Mobile number already exists.")

        email = request.data.get("email", staff.email)

        if email and Staff.objects.filter(school=school, email=email).exclude(id=staff.id).exists():
            application_logger.warning(
                "staff_update_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=str(staff.id),
                reason="email_already_exists",
            )
            return CustomResponse.errorResponse(description="Email already exists.")

        staff_type = request.data.get("staff_type", staff.staff_type)

        if staff_type not in Staff.StaffType.values:
            application_logger.warning(
                "staff_update_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=str(staff.id),
                staff_type=staff_type,
                reason="invalid_staff_type",
            )
            return CustomResponse.errorResponse(description="Invalid staff type.")

        gender = request.data.get("gender", staff.gender)

        if gender not in Staff.Gender.values:
            application_logger.warning(
                "staff_update_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=str(staff.id),
                gender=gender,
                reason="invalid_gender",
            )
            return CustomResponse.errorResponse(description="Invalid gender.")

        role = Roles.objects.filter(role_name=staff_type).first()

        if role is None:
            application_logger.warning(
                "staff_update_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=str(staff.id),
                staff_type=staff_type,
                reason="role_not_configured",
            )
            return CustomResponse.errorResponse(description=f"{staff_type.title()} role not configured.")

        try:

            with transaction.atomic():

                user = staff.user

                user.first_name = request.data.get("name", staff.name)
                user.mobile = mobile
                user.email = email
                user.gender = gender
                user.date_of_birth = request.data.get("date_of_birth", staff.date_of_birth)
                user.profile_image = request.data.get("profile_image", staff.profile_image)
                user.save()

                staff.branch = branch
                staff.employee_id = employee_id
                staff.staff_type = staff_type
                staff.name = request.data.get("name", staff.name)
                staff.gender = gender
                staff.date_of_birth = request.data.get("date_of_birth", staff.date_of_birth)
                staff.mobile = mobile
                staff.email = email
                staff.qualification = request.data.get("qualification", staff.qualification)
                staff.experience = request.data.get("experience", staff.experience)
                staff.joining_date = request.data.get("joining_date", staff.joining_date)
                staff.status = request.data.get("status", staff.status)
                staff.profile_image = request.data.get("profile_image", staff.profile_image)
                staff.address = request.data.get("address", staff.address)
                staff.emergency_contact_name = request.data.get("emergency_contact_name", staff.emergency_contact_name)
                staff.emergency_contact_mobile = request.data.get("emergency_contact_mobile", staff.emergency_contact_mobile)
                staff.save()

                UserRoles.objects.filter(user=user, school=school).update(role=role)

        except Exception as e:
            application_logger.exception(
                "staff_update_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=str(staff.id),
                reason="staff_update_failed",
                error=str(e),
            )
            return CustomResponse.errorResponse(description=str(e))

        application_logger.info(
            "staff_updated",
            requested_by=str(request.user.id),
            school_id=str(school.id),
            staff_id=str(staff.id),
            employee_id=staff.employee_id,
            staff_type=staff.staff_type,
            branch_id=str(staff.branch_id) if staff.branch_id else None,
            role=role.role_name,
        )

        return CustomResponse.successResponse(
            description="Staff updated successfully.",
            data={
                "id": str(staff.id),
                "employee_id": staff.employee_id,
                "name": staff.name,
                "role": role.role_name,
            },
        )


class CreateStaffDocumentAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "staff.document.create"

    def post(self, request):

        school = request.school
        staff_id = request.data.get("staff_id")
        document_type = request.data.get("document_type")

        application_logger.info(
            "staff_document_create_requested",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            staff_id=staff_id,
            document_type=document_type,
        )

        if school is None:
            application_logger.warning(
                "staff_document_create_failed",
                requested_by=str(request.user.id),
                reason="school_not_found",
            )
            return CustomResponse.errorResponse(description="School not found.")

        required_fields = ["staff_id", "document_type", "document_name", "document_url"]

        for field in required_fields:
            if request.data.get(field) in [None, ""]:
                application_logger.warning(
                    "staff_document_create_failed",
                    requested_by=str(request.user.id),
                    school_id=str(school.id),
                    staff_id=staff_id,
                    field=field,
                    reason="required_field_missing",
                )
                return CustomResponse.errorResponse(description=f"{field} is required.")

        staff = Staff.objects.filter(id=staff_id, school=school).first()

        if staff is None:
            application_logger.warning(
                "staff_document_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=staff_id,
                reason="staff_not_found",
            )
            return CustomResponse.errorResponse(description="Staff not found.")

        if document_type not in StaffDocument.DocumentType.values:
            application_logger.warning(
                "staff_document_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=str(staff.id),
                document_type=document_type,
                reason="invalid_document_type",
            )
            return CustomResponse.errorResponse(description="Invalid document type.")

        if StaffDocument.objects.filter(staff=staff, document_type=document_type).exists():
            application_logger.warning(
                "staff_document_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=str(staff.id),
                document_type=document_type,
                reason="document_already_exists",
            )
            return CustomResponse.errorResponse(
                description=f"{document_type.replace('_', ' ').title()} already exists."
            )

        try:

            with transaction.atomic():

                document = StaffDocument.objects.create(
                    staff=staff,
                    document_type=document_type,
                    document_name=request.data.get("document_name").strip(),
                    document_number=request.data.get("document_number"),
                    document_url=request.data.get("document_url"),
                    issue_date=request.data.get("issue_date"),
                    expiry_date=request.data.get("expiry_date"),
                    is_verified=request.data.get("is_verified", False),
                    remarks=request.data.get("remarks"),
                )

        except Exception as e:
            application_logger.exception(
                "staff_document_create_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=str(staff.id),
                document_type=document_type,
                reason="staff_document_creation_failed",
                error=str(e),
            )
            return CustomResponse.errorResponse(description=str(e))

        application_logger.info(
            "staff_document_created",
            requested_by=str(request.user.id),
            school_id=str(school.id),
            staff_id=str(staff.id),
            document_id=str(document.id),
            document_type=document.document_type,
            is_verified=document.is_verified,
        )

        return CustomResponse.successResponse(
            description="Document uploaded successfully.",
            data={
                "id": str(document.id),
                "staff_id": str(staff.id),
                "document_type": document.document_type,
                "document_name": document.document_name,
                "document_url": document.document_url,
                "is_verified": document.is_verified,
            },
        )

class StaffDocumentListAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "staff.document.view"

    def get(self, request):

        school = request.school
        staff_id = request.GET.get("staff_id")

        application_logger.info(
            "staff_document_list_requested",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            staff_id=staff_id,
        )

        if school is None:

            application_logger.warning(
                "staff_document_list_failed",
                requested_by=str(request.user.id),
                staff_id=staff_id,
                reason="school_not_found",
            )

            return CustomResponse.errorResponse(
                description="School not found."
            )

        if not staff_id:

            application_logger.warning(
                "staff_document_list_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                reason="staff_id_required",
            )

            return CustomResponse.errorResponse(
                description="staff_id is required."
            )

        try:

            staff = Staff.objects.filter(
                id=staff_id,
                school=school,
            ).first()

            if staff is None:

                application_logger.warning(
                    "staff_document_list_failed",
                    requested_by=str(request.user.id),
                    school_id=str(school.id),
                    staff_id=staff_id,
                    reason="staff_not_found",
                )

                return CustomResponse.errorResponse(
                    description="Staff not found."
                )

            documents = StaffDocument.objects.filter(
                staff=staff
            ).order_by("document_type")

            data = []

            for document in documents:

                data.append({
                    "id": str(document.id),
                    "document_type": document.document_type,
                    "document_type_display": document.get_document_type_display(),
                    "document_name": document.document_name,
                    "document_number": document.document_number,
                    "document_url": document.document_url,
                    "issue_date": document.issue_date,
                    "expiry_date": document.expiry_date,
                    "is_verified": document.is_verified,
                    "remarks": document.remarks,
                    "created_at": document.created_at,
                    "updated_at": document.updated_at,
                })

        except Exception as e:

            application_logger.exception(
                "staff_document_list_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                staff_id=staff_id,
                reason="staff_document_list_fetch_failed",
                error=str(e),
            )

            return CustomResponse.errorResponse(
                description=str(e)
            )

        application_logger.info(
            "staff_document_list_fetched",
            requested_by=str(request.user.id),
            school_id=str(school.id),
            staff_id=str(staff.id),
            returned_count=len(data),
        )

        return CustomResponse.successResponse(
            description="Staff documents fetched successfully.",
            data={
                "staff_id": str(staff.id),
                "staff_name": staff.name,
                "documents": data,
            },
        )


class UpdateStaffDocumentAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]
    required_permission = "staff.document.update"

    def put(self, request, document_id):

        school = request.school

        application_logger.info(
            "staff_document_update_requested",
            requested_by=str(request.user.id),
            school_id=str(school.id) if school else None,
            document_id=str(document_id),
        )

        if school is None:

            application_logger.warning(
                "staff_document_update_failed",
                requested_by=str(request.user.id),
                document_id=str(document_id),
                reason="school_not_found",
            )

            return CustomResponse.errorResponse(
                description="School not found."
            )

        document = StaffDocument.objects.select_related(
            "staff"
        ).filter(
            id=document_id,
            staff__school=school,
        ).first()

        if document is None:

            application_logger.warning(
                "staff_document_update_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                document_id=str(document_id),
                reason="document_not_found",
            )

            return CustomResponse.errorResponse(
                description="Document not found."
            )

        document_type = request.data.get("document_type")

        if document_type:

            if document_type not in StaffDocument.DocumentType.values:

                application_logger.warning(
                    "staff_document_update_failed",
                    requested_by=str(request.user.id),
                    school_id=str(school.id),
                    document_id=str(document.id),
                    staff_id=str(document.staff.id),
                    document_type=document_type,
                    reason="invalid_document_type",
                )

                return CustomResponse.errorResponse(
                    description="Invalid document type."
                )

            if StaffDocument.objects.filter(
                staff=document.staff,
                document_type=document_type,
            ).exclude(
                id=document.id
            ).exists():

                application_logger.warning(
                    "staff_document_update_failed",
                    requested_by=str(request.user.id),
                    school_id=str(school.id),
                    document_id=str(document.id),
                    staff_id=str(document.staff.id),
                    document_type=document_type,
                    reason="document_type_already_exists",
                )

                return CustomResponse.errorResponse(
                    description=f"{document_type.replace('_', ' ').title()} already exists."
                )

            document.document_type = document_type

        if request.data.get("document_name") not in [None, ""]:
            document.document_name = request.data.get("document_name").strip()

        if "document_number" in request.data:
            document.document_number = request.data.get("document_number")

        if request.data.get("document_url") not in [None, ""]:
            document.document_url = request.data.get("document_url")

        if "issue_date" in request.data:
            document.issue_date = request.data.get("issue_date") or None

        if "expiry_date" in request.data:
            document.expiry_date = request.data.get("expiry_date") or None

        if "is_verified" in request.data:
            document.is_verified = request.data.get("is_verified")

        if "remarks" in request.data:
            document.remarks = request.data.get("remarks")

        try:

            with transaction.atomic():
                document.save()

        except Exception as e:

            application_logger.exception(
                "staff_document_update_failed",
                requested_by=str(request.user.id),
                school_id=str(school.id),
                document_id=str(document.id),
                staff_id=str(document.staff.id),
                reason="document_update_failed",
                error=str(e),
            )

            return CustomResponse.errorResponse(
                description=str(e)
            )

        application_logger.info(
            "staff_document_updated",
            requested_by=str(request.user.id),
            school_id=str(school.id),
            document_id=str(document.id),
            staff_id=str(document.staff.id),
            document_type=document.document_type,
        )

        return CustomResponse.successResponse(
            description="Document updated successfully.",
            data={
                "id": str(document.id),
                "staff_id": str(document.staff.id),
                "document_type": document.document_type,
                "document_name": document.document_name,
                "document_number": document.document_number,
                "document_url": document.document_url,
                "issue_date": document.issue_date,
                "expiry_date": document.expiry_date,
                "is_verified": document.is_verified,
                "remarks": document.remarks,
            },
        )

class CreateSubjectAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]

    required_permission = "subject.create"

    def post(self, request):

        school = request.school

        application_logger.info(
            "subject_create_started",
            user_id=str(request.user.id),
            school_id=str(school.id) if school else None,
        )

        try:

            if school is None:

                application_logger.warning(
                    "subject_create_failed",
                    reason="school_not_found",
                    user_id=str(request.user.id),
                )

                return CustomResponse.errorResponse(
                    description="School not found."
                )

            academic_year_id = request.data.get("academic_year_id")
            name = request.data.get("name")

            if not academic_year_id:

                return CustomResponse.errorResponse(
                    description="academic_year_id is required."
                )

            if not name:

                return CustomResponse.errorResponse(
                    description="name is required."
                )

            academic_year = AcademicYear.objects.filter(
                id=academic_year_id,
                school=school,
            ).first()

            if academic_year is None:

                application_logger.warning(
                    "subject_create_failed",
                    school_id=str(school.id),
                    academic_year_id=str(academic_year_id),
                    reason="academic_year_not_found",
                )

                return CustomResponse.errorResponse(
                    description="Academic year not found."
                )

            if Subject.objects.filter(
                school=school,
                academic_year=academic_year,
                name__iexact=name.strip(),
            ).exists():

                application_logger.warning(
                    "subject_create_failed",
                    school_id=str(school.id),
                    academic_year_id=str(academic_year.id),
                    subject_name=name,
                    reason="subject_already_exists",
                )

                return CustomResponse.errorResponse(
                    description="Subject already exists."
                )

            with transaction.atomic():

                subject = Subject.objects.create(
                    school=school,
                    academic_year=academic_year,
                    name=name.strip(),
                    description=request.data.get("description"),
                    status=request.data.get(
                        "status",
                        Subject.Status.ACTIVE,
                    ),
                )

            application_logger.info(
                "subject_created",
                user_id=str(request.user.id),
                school_id=str(school.id),
                academic_year_id=str(academic_year.id),
                subject_id=str(subject.id),
            )

            return CustomResponse.successResponse(
                description="Subject created successfully.",
                data={
                    "id": str(subject.id),
                    "name": subject.name,
                    "description": subject.description,
                    "status": subject.status,
                    "academic_year": {
                        "id": str(academic_year.id),
                        "name": academic_year.name,
                    },
                },
            )

        except Exception:

            application_logger.exception(
                "subject_create_failed",
                user_id=str(request.user.id),
                school_id=str(school.id) if school else None,
            )

            return CustomResponse.errorResponse(
                description="Something went wrong while creating subject."
            )




class SubjectListAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]

    required_permission = "subject.view"

    def get(self, request):

        school = request.school

        application_logger.info(
            "subject_list_started",
            user_id=str(request.user.id),
            school_id=str(school.id) if school else None,
        )

        try:

            if school is None:

                application_logger.warning(
                    "subject_list_failed",
                    reason="school_not_found",
                    user_id=str(request.user.id),
                )

                return CustomResponse.errorResponse(
                    description="School not found."
                )

            subjects = Subject.objects.select_related(
                "academic_year",
            ).filter(
                school=school,
            )

            academic_year_id = request.query_params.get(
                "academic_year_id"
            )

            status = request.query_params.get(
                "status"
            )

            search = request.query_params.get(
                "search"
            )

            if academic_year_id:

                subjects = subjects.filter(
                    academic_year_id=academic_year_id,
                )

            if status:

                subjects = subjects.filter(
                    status=status,
                )

            if search:

                subjects = subjects.filter(
                    name__icontains=search.strip(),
                )

            subjects = subjects.order_by(
                "name",
            )

            data = []

            for subject in subjects:

                data.append({
                    "id": str(subject.id),
                    "name": subject.name,
                    "description": subject.description,
                    "status": subject.status,
                    "academic_year": {
                        "id": str(subject.academic_year.id),
                        "name": subject.academic_year.name,
                    },
                })

            application_logger.info(
                "subject_list_fetched",
                user_id=str(request.user.id),
                school_id=str(school.id),
                total_count=len(data),
            )

            return CustomResponse.successResponse(
                description="Subjects fetched successfully.",
                data=data,
                total_count=len(data),
            )

        except Exception:

            application_logger.exception(
                "subject_list_failed",
                user_id=str(request.user.id),
                school_id=str(school.id) if school else None,
            )

            return CustomResponse.errorResponse(
                description="Something went wrong while fetching subjects."
            )

class SubjectUpdateAPIView(APIView):

    permission_classes = [IsAuthenticated, HasPermission]

    required_permission = "subject.update"

    def put(self, request, subject_id):

        school = request.school

        application_logger.info(
            "subject_update_started",
            user_id=str(request.user.id),
            school_id=str(school.id) if school else None,
            subject_id=str(subject_id),
        )

        try:

            if school is None:

                application_logger.warning(
                    "subject_update_failed",
                    reason="school_not_found",
                    user_id=str(request.user.id),
                )

                return CustomResponse.errorResponse(
                    description="School not found."
                )

            subject = Subject.objects.select_related(
                "academic_year",
            ).filter(
                id=subject_id,
                school=school,
            ).first()

            if subject is None:

                application_logger.warning(
                    "subject_update_failed",
                    reason="subject_not_found",
                    school_id=str(school.id),
                    subject_id=str(subject_id),
                )

                return CustomResponse.errorResponse(
                    description="Subject not found."
                )

            academic_year_id = request.data.get("academic_year_id")
            name = request.data.get("name")

            if not academic_year_id:

                return CustomResponse.errorResponse(
                    description="academic_year_id is required."
                )

            if not name:

                return CustomResponse.errorResponse(
                    description="name is required."
                )

            academic_year = AcademicYear.objects.filter(
                id=academic_year_id,
                school=school,
            ).first()

            if academic_year is None:

                application_logger.warning(
                    "subject_update_failed",
                    reason="academic_year_not_found",
                    academic_year_id=str(academic_year_id),
                )

                return CustomResponse.errorResponse(
                    description="Academic year not found."
                )

            if Subject.objects.filter(
                school=school,
                academic_year=academic_year,
                name__iexact=name.strip(),
            ).exclude(
                id=subject.id,
            ).exists():

                application_logger.warning(
                    "subject_update_failed",
                    reason="subject_already_exists",
                    subject_name=name,
                )

                return CustomResponse.errorResponse(
                    description="Subject already exists."
                )

            status = request.data.get(
                "status",
                subject.status,
            )

            if status not in Subject.Status.values:

                return CustomResponse.errorResponse(
                    description="Invalid status."
                )

            with transaction.atomic():

                subject.academic_year = academic_year
                subject.name = name.strip()
                subject.description = request.data.get("description")
                subject.status = status

                subject.save()

            application_logger.info(
                "subject_updated",
                user_id=str(request.user.id),
                school_id=str(school.id),
                subject_id=str(subject.id),
            )

            return CustomResponse.successResponse(
                description="Subject updated successfully.",
                data={
                    "id": str(subject.id),
                    "name": subject.name,
                    "description": subject.description,
                    "status": subject.status,
                    "academic_year": {
                        "id": str(subject.academic_year.id),
                        "name": subject.academic_year.name,
                    },
                },
            )

        except Exception:

            application_logger.exception(
                "subject_update_failed",
                user_id=str(request.user.id),
                school_id=str(school.id) if school else None,
                subject_id=str(subject_id),
            )

            return CustomResponse.errorResponse(
                description="Something went wrong while updating subject."
            )