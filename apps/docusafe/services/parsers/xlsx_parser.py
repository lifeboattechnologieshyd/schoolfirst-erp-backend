"""
XLSX parser.

Handles: .xlsx, .xls
Uses openpyxl to iterate sheets and rows, producing
sheet-aware DocumentBlocks with page_index = sheet index.
"""

from __future__ import annotations

import io

import structlog
from django.core.files.storage import default_storage
from openpyxl import load_workbook

from apps.docusafe.services.parsers.document_block import DocumentBlock

logger = structlog.getLogger("default")


def parse_xlsx(file_path: str) -> list[DocumentBlock]:
    """
    Parse an Excel file into sheet-aware DocumentBlocks.

    Each sheet is treated as a separate page (page_index = sheet index).
    The first non-empty row of each sheet is treated as a header.
    Subsequent rows become individual sheet_row blocks.
    """
    wb = _load_workbook_from_s3(file_path)
    blocks: list[DocumentBlock] = []

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # Find the first non-empty row as header
        header_row = None
        data_start = 0
        for i, row in enumerate(rows):
            if any(cell is not None for cell in row):
                header_row = [str(cell) if cell is not None else "" for cell in row]
                data_start = i + 1
                break

        if header_row is None:
            continue

        # Emit header as a heading block (sets the section)
        header_text = " | ".join(header_row)
        blocks.append(
            DocumentBlock(
                text=header_text,
                block_type="heading",
                page_index=sheet_idx,
                section=sheet_name,
                metadata={"sheet_name": sheet_name, "heading_level": 1},
            )
        )

        # Emit each data row
        for row_idx, row in enumerate(rows[data_start:], start=data_start):
            cells = [str(cell) if cell is not None else "" for cell in row]

            # Skip completely empty rows
            if not any(cell.strip() for cell in cells):
                continue

            # Format as "header1: value1 | header2: value2 | ..."
            if len(cells) == len(header_row):
                row_text = " | ".join(f"{h}: {v}" for h, v in zip(header_row, cells, strict=True))
            else:
                row_text = " | ".join(cells)

            blocks.append(
                DocumentBlock(
                    text=row_text,
                    block_type="sheet_row",
                    page_index=sheet_idx,
                    section=sheet_name,
                    metadata={"sheet_name": sheet_name, "row_index": row_idx},
                )
            )

    wb.close()
    return blocks


def _load_workbook_from_s3(file_path: str):
    """Download an Excel file from S3 and return an openpyxl workbook."""
    try:
        with default_storage.open(file_path, "rb") as file_obj:
            content = file_obj.read()
        return load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        logger.exception("Failed to load Excel from S3", file_path=file_path)
        raise
